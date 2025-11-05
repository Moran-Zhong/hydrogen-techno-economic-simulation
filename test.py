#!/usr/bin/env python3
"""
Test configuration file transfer functionality for dual-location wind power model
"""

# Ensure loading the correct HOPP version

import pandas as pd
import numpy as np
# Import HOPP modules
from hopp.optimization.system_optimizer import SystemOptimizer
from hopp.tools.analysis import EconomicCalculator
from hopp.utilities import ConfigManager
from hopp.utilities.keys import set_developer_nrel_gov_key
from hopp.simulation.resource_files import ResourceDataManager
import gc
import os
from openpyxl import load_workbook

# Set your NREL API key (required for downloading solar data)
set_developer_nrel_gov_key('uZQBIA4mfVd7kj9ofLFNY0rmEVy5qx1v2fBpTAIG')


def write_result_to_excel(file_path, target_demand_met, ratio, result):
    """Write single result to Excel file"""

    # Create worksheet name
    sheet_name = f"demand_{target_demand_met}"
    sheet_name = sheet_name[:31]  # Excel worksheet name limit

    # Check if file exists
    if os.path.exists(file_path):
        # Open existing file
        book = load_workbook(file_path)

        # Check if worksheet already exists
        if sheet_name in book.sheetnames:
            # If worksheet exists, read existing data
            df_existing = pd.read_excel(file_path, sheet_name=sheet_name)

            # Add new column
            new_column_index = len(df_existing.columns)

            # Add ratio value to first row
            if new_column_index == 1:  # Only Metric column
                df_existing.loc[0, f"Case_{new_column_index}"] = ratio
            else:
                df_existing.loc[0, f"Case_{new_column_index}"] = ratio

            # Add other metrics
            for idx, (metric, value) in enumerate(result.items()):
                # Find metric's row position in DataFrame
                row_idx = df_existing[df_existing['Metric'] == metric].index
                if len(row_idx) > 0:
                    # If metric exists, update corresponding column value
                    df_existing.loc[row_idx[0], f"Case_{new_column_index}"] = value
                else:
                    # If metric doesn't exist, add new row
                    new_row = pd.Series([metric] + [None] * (new_column_index - 1) + [value],
                                        index=['Metric'] + [f"Case_{i + 1}" for i in range(new_column_index)])
                    df_existing = pd.concat([df_existing, new_row.to_frame().T], ignore_index=True)

            # Save updated worksheet
            with pd.ExcelWriter(file_path, engine="openpyxl", mode='a', if_sheet_exists='replace') as writer:
                df_existing.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # If worksheet doesn't exist, create new worksheet
            all_metrics = list(result.keys())

            # Create DataFrame
            rows = []

            # First row: ratio value
            ratio_row = {"Metric": "Ratio", "Case_1": ratio}
            rows.append(ratio_row)

            # Subsequent rows: various metrics (maintain original order of result dictionary)
            for metric in all_metrics:
                metric_row = {"Metric": metric, "Case_1": result.get(metric, "N/A")}
                rows.append(metric_row)

            df = pd.DataFrame(rows)

            # Add to existing workbook
            with pd.ExcelWriter(file_path, engine="openpyxl", mode='a', if_sheet_exists='new') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # If file doesn't exist, create new file and worksheet
        all_metrics = list(result.keys())

        # Create DataFrame
        rows = []

        # First row: ratio value
        ratio_row = {"Metric": "Ratio", "Case_1": ratio}
        rows.append(ratio_row)

        # Subsequent rows: various metrics (maintain original order of result dictionary)
        for metric in all_metrics:
            metric_row = {"Metric": metric, "Case_1": result.get(metric, "N/A")}
            rows.append(metric_row)

        df = pd.DataFrame(rows)

        # Create new file
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)


# Initialize resource manager for downloading data
resource_manager = ResourceDataManager(
    api_key='uZQBIA4mfVd7kj9ofLFNY0rmEVy5qx1v2fBpTAIG',
    email='z5536449@ad.unsw.edu.au'
)

# Location information for solar
solar_latitude = -19.5825
solar_longitude = 146.8398

# Location information for wind (different location)
wind_latitude1 = -26.73274363843257
wind_longitude1 = 151.4722470563812

wind_latitude2 = -17.183333
wind_longitude2 = 145.366666 #145.06

# Download resource data
solar_path = resource_manager.download_solar_data(
    latitude=solar_latitude,
    longitude=solar_longitude,
    year="2016"
)
wind_path1 = resource_manager.download_wind_data(
    latitude=wind_latitude1,
    longitude=wind_longitude1,
    start_date="20160101",
    end_date="20161231",
)
wind_path2 = resource_manager.download_wind_data(
    latitude=wind_latitude2,
    longitude=wind_longitude2,
    start_date="20160101",
    end_date="20161231"
)

wind_path = [wind_path1, wind_path2]

# read load data
csv_path = r'D:\python\Multi-wind simulation\Load data.csv'
csv_df = pd.read_csv(csv_path)
load_data = csv_df.iloc[:, 0].tolist()

# Use configuration file in workspace
yaml_file_path = str(r"D:\python\Multi-wind simulation\examples\inputs\test-separated-simple.yaml")
print(f"Using configuration file: {yaml_file_path}")
config = ConfigManager.load_yaml_safely(yaml_file_path)  # Use static method of class

# Update configuration with location and resource files
config['site']['data']['lat'] = solar_latitude
config['site']['data']['lon'] = solar_longitude
config['site']['solar_resource_file'] = solar_path
config['site']['wind_resource_files'] = wind_path
config['site']['desired_schedule'] = load_data

ConfigManager.save_yaml_safely(config, yaml_file_path)

# ==== External parameter settings ====
target_demand_met = 85.0  # Target demand met (time_load_met) percentage
demand_met_tolerance = 0.1  # Tolerance range (percentage)
# =====================================

# Initialize Excel file
file_path = "output.xlsx"
if os.path.exists(file_path):
    os.remove(file_path)  # Delete existing file

# Main loop
for target_demand_met in np.arange(97, 98, 2):  # 85% to 99%, interval 2%
    target_demand_met = round(target_demand_met, 1)  # Ensure precision
    print(f"\n{'=' * 60}")
    print(f"Starting optimization target_demand_met = {target_demand_met}%")
    print(f"{'=' * 60}")

    # Loop through different wind farm ratios
    case_count = 0  # Track number of cases under current target_demand_met
    for i in np.arange(0.5, 0.85, 0.5):  # 0 to 1, step 0.1
        wind1_ratio = round(i, 4)  # Round to avoid floating point precision issues
        wind2_ratio = round(1 - wind1_ratio, 4)

        print(f"\n===== Running ratio: Wind Farm 1 = {wind1_ratio:.1f}, Wind Farm 2 = {wind2_ratio:.1f} =====")
        # Avoid division by zero
        if wind1_ratio == 0:
            ratio = 100  # A large number indicating Wind Farm 1 completely dominates
        if wind2_ratio == 0:
            ratio = 0.001
        else:
            ratio = wind2_ratio / wind1_ratio

        # Set ratio and run optimization

        # Initialize components for optimization
        economic_calculator = EconomicCalculator(
            discount_rate=0.0588,
            project_lifetime=25
        )

        optimizer = SystemOptimizer(
            yaml_file_path=yaml_file_path,
            economic_calculator=economic_calculator,
            enable_flexible_load=False, #set to False to disable flexible load
            max_load_reduction_percentage=0  # 20% maximum load reduction
        )

        # Set demand met target constraint
        optimizer.set_demand_met_target(target_demand_met, demand_met_tolerance)

        # Define optimization bounds
        bounds = [
            (120000, 400000),  # PV capacity (kW)
            (35, 65),  # Wind turbines (5MW each)
            (240000, 400000),  # Battery capacity (kWh)
            (5000, 30000),  # Battery power (kW)
        ]

        # Define initial conditions (10% of range)
        initial_conditions = [
            [bound[0] + (bound[1] - bound[0]) * 0.5 for bound in bounds],
            [bound[0] + (bound[1] - bound[0]) * 0.2 for bound in bounds],
            [bound[0] + (bound[1] - bound[0]) * 0.8 for bound in bounds],
        ]

        # Set ratio and run optimization
        optimizer.set_turbine_ratio(ratio)

        # Run optimization
        result = optimizer.optimize_system(bounds, initial_conditions)
        # result = optimizer.optimize_system_de(bounds)
        # result = optimizer.optimize_system_ga(bounds)

        # Print results
        if result:
            print("\nOptimization Results:")
            print(f"PV Capacity: {result['PV Capacity (kW)']:.2f} kW")
            print(f"Wind Turbines: {result['Wind Turbine Capacity (kW)'] / 5000:.0f} x 5MW")
            print(f"Battery Capacity: {result['Battery Energy Capacity (kWh)']:.2f} kWh")
            print(f"Battery Power: {result['Battery Power Capacity (kW)']:.2f} kW")
            print(f"\nLCOE: ${result['System LCOE (cents/kWh)']:.4f}cents/kWh")
            print(f"Demand Met: {result['Time Load Met (%)']:.2f}%")
            print(f"Target Demand Met: {target_demand_met}% ± {demand_met_tolerance}%")

            # Write to Excel immediately
            write_result_to_excel(file_path, target_demand_met, ratio, result)
            del result

        gc.collect()
    gc.collect()
    print(f"\nCompleted all ratio optimizations for target_demand_met = {target_demand_met}%")

# Final memory cleanup
gc.collect()

print(f"\nAll results written incrementally to {file_path}")
